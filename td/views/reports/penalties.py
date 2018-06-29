from django.shortcuts import render, redirect
from django.http import HttpResponse


def penalties_report_form(request):
    from td.forms import PeriodForm
    from datetime import date

    form = PeriodForm()

    context = {'form': form}
    return render(request, 'td/reports/penalties_form.html', context)


def penalties_report(request, xls=False):
    from td.models import Penalty
    from td.forms import PeriodForm
    from datetime import datetime

    if request.method != 'POST':
        return redirect(to='td:penalties_report_form')

    form = PeriodForm(request.POST)
    if form.data['start_date']:
        start_date = datetime.strptime(form.data['start_date'], '%d.%m.%Y')
    else:
        start_date = datetime(2000, 1, 1, 0, 0)
    if form.data['end_date']:
        end_date = datetime.strptime(form.data['end_date'], '%d.%m.%Y')
    else:
        end_date = datetime(2099, 12, 31, 0, 0)

    penalties = Penalty.objects\
        .filter(car__active=True, date__gte=start_date, date__lte=end_date)\
        .order_by('car__name', 'car__number', 'date')

    if not xls:
        context = {'penalties': penalties}
        return render(request, 'td/reports/penalties.html', context)
    else:
        from main import settings
        from django.db.models import Sum
        import os
        import openpyxl
        from openpyxl.styles import Alignment
        from openpyxl.writer.excel import save_virtual_workbook

        wb = openpyxl.load_workbook(os.path.join(settings.TEMPLATES[0]['DIRS'][0], 'td', 'reports', 'penalties.xlsx'))
        ws = wb.active
        wrap_alignment = Alignment(wrap_text=True)

        row = 5
        current_car = None

        for n, p in enumerate(penalties):
            if current_car != p.car_id:
                ws.cell(column=1, row=row, value=p.car.name + ' ' + p.car.number)
                ws.cell(column=7, row=row, value='Общая сумма:')
                total_amount = penalties.filter(car_id=p.car_id).aggregate(Sum('amount'))['amount__sum']
                ws.cell(column=8, row=row, value=total_amount)
                ws.cell(column=9, row=row, value='Оплачено:')
                payed_count = penalties.filter(car_id=p.car_id, payed=True).count()
                total_count = penalties.filter(car_id=p.car_id).count()
                ws.cell(column=10, row=row, value=str(payed_count)+'/'+str(total_count))
                for c in range(1, 12):
                    ws.cell(column=c, row=row).style = 'car_header'
                ws.merge_cells(start_column=1, end_column=6, start_row=row, end_row=row)
                current_car = p.car_id
                row += 1

            ws.cell(column=1, row=row, value=n+1)
            ws.cell(column=2, row=row, value=p.car.name)
            ws.cell(column=3, row=row, value=p.car.number)
            ws.cell(column=4, row=row, value=p.number)
            ws.cell(column=5, row=row, value=p.car.get_ownership_type_display())
            if p.date is not None:
                ws.cell(column=6, row=row, value=p.date.strftime('%d.%m.%Y'))
            ws.cell(column=7, row=row, value=p.description)
            ws.cell(column=8, row=row, value=p.amount)
            if p.obtain_date is not None:
                ws.cell(column=9, row=row, value=p.obtain_date.strftime('%d.%m.%Y'))
            if p.payed:
                state = 'Оплачен'
            else:
                state = 'Не оплачен'
            ws.cell(column=10, row=row, value=state)
            if p.pay_date is not None:
                ws.cell(column=11, row=row, value=p.pay_date.strftime('%d.%m.%Y'))
            for c in range(1,12):
                ws.cell(column=c, row=row).style = 'cell'
                ws.cell(column=c, row=row).alignment = wrap_alignment
            row += 1

        # ------- totals -------

        ws.merge_cells(start_column=1, end_column=6, start_row=row, end_row=row)
        ws.cell(column=1, row=row, value='ИТОГО')
        ws.cell(column=7, row=row, value='Общая сумма:')
        total_amount = penalties.aggregate(Sum('amount'))['amount__sum']
        ws.cell(column=8, row=row, value=total_amount)
        ws.cell(column=9, row=row, value='Оплачено:')
        payed_count = penalties.filter(payed=True).count()
        total_count = penalties.count()
        ws.cell(column=10, row=row, value=str(payed_count) + '/' + str(total_count))
        for c in range(1, 12):
            ws.cell(column=c, row=row).style = 'car_header'
        row += 1

        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=report.xlsx"

        return response
