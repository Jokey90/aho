from django.shortcuts import render, redirect
from django.http import HttpResponse


def mileage_report_form(request):
    from td.forms import YearSelectForm
    from datetime import date

    form = YearSelectForm()

    context = {'form': form}
    return render(request, 'td/reports/mileage_form.html', context)


def mileage_report(request, xls=False):
    from td.models import Mileage
    from td.forms import YearSelectForm
    from datetime import datetime

    if request.method != 'POST':
        return redirect(to='td:mileage_report_form')

    form = YearSelectForm(request.POST)
    year = int(form.data['year'])

    mileages = Mileage.objects.with_periods().filter(year=year).order_by('car__name', 'car__number', 'month').all()

    if not xls:
        context = {
            'mileages': mileages,
            'year': year
        }
        return render(request, 'td/reports/mileage.html', context)
    else:
        from main import settings
        import os
        import openpyxl
        from openpyxl.styles import Alignment
        from openpyxl.writer.excel import save_virtual_workbook

        wb = openpyxl.load_workbook(os.path.join(settings.TEMPLATES[0]['DIRS'][0], 'td', 'reports', 'mileage.xlsx'))
        ws = wb.active
        wrap_alignment = Alignment(wrap_text=True)

        mileages_by_car = []
        current_car = None
        temp_array = []
        for elem in mileages:
            if elem.car_id != current_car:
                if len(temp_array) > 0:
                    mileages_by_car.append(temp_array)
                    temp_array = []
                current_car = elem.car_id
            temp_array.append(elem)
        mileages_by_car.append(temp_array)

        row = 4
        ws.cell(column=3, row=1, value='Отчет по затратам на топливо относительно пройденных км за '+str(year)+' год')
        for n, car_mileages in enumerate(mileages_by_car):
            ws.merge_cells(start_column=1, end_column=1, start_row=row, end_row=row+1)
            ws.merge_cells(start_column=2, end_column=2, start_row=row, end_row=row+1)
            ws.merge_cells(start_column=3, end_column=3, start_row=row, end_row=row+1)
            ws.merge_cells(start_column=4, end_column=4, start_row=row, end_row=row+1)
            ws.merge_cells(start_column=5, end_column=5, start_row=row, end_row=row+1)
            ws.cell(column=1, row=row, value=n+1).style = 'centered_cell'
            ws.cell(column=2, row=row, value=car_mileages[0].car.name)
            ws.cell(column=3, row=row, value=car_mileages[0].car.number)
            ws.cell(column=4, row=row, value=car_mileages[0].car.vin)
            ws.cell(column=5, row=row, value=car_mileages[0].car.release_year)
            ws.cell(column=6, row=row, value='Пробег по одометру')
            ws.cell(column=6, row=row+1, value='Сумма расходов руб./км')
            for m in range(1,13):
                ws.cell(column=6 + m, row=row, value=0)
                ws.cell(column=6 + m, row=row + 1, value=0)
            for mileage in car_mileages:
                ws.cell(column=6+mileage.month, row=row, value=mileage.diff())
                ws.cell(column=6+mileage.month, row=row+1, value=mileage.gasoline_rate())
            row += 2

        for r in range(2, row):
            for c in range(1, 19):
                ws.cell(column=c, row=r).alignment = wrap_alignment
                ws.cell(column=c, row=r).style = 'centered_cell'

        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=report.xlsx"

        return response
