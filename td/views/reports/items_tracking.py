from django.shortcuts import render, redirect
from django.http import HttpResponse


def tracking_report_form(request):
    from td.forms import CarDriverPeriodForm

    form = CarDriverPeriodForm()
    context = {'form': form}
    return render(request, 'td/reports/tracking_form.html', context)


def tracking_report(request, xls=False):
    from td.forms import CarDriverPeriodForm
    from td.models import ItemTracking, ProxyTracking
    from datetime import datetime

    if request.method != 'POST':
        return redirect(to='td:tracking_report_form')

    form = CarDriverPeriodForm(request.POST)

    if form.data['start_date']:
        start_date = datetime.strptime(form.data['start_date'], '%d.%m.%Y')
    else:
        start_date = datetime(2000, 1, 1, 0, 0)
    if form.data['end_date']:
        end_date = datetime.strptime(form.data['end_date'], '%d.%m.%Y')
    else:
        end_date = datetime(2099, 12, 31, 0, 0)
    driver_id = int(form.data['driver'])
    car_id = int(form.data['car'])

    item_tracking_list = ItemTracking.objects.filter(date__gte=start_date, date__lte=end_date, item__active=True).order_by('date', 'id')
    proxy_tracking_list = ProxyTracking.objects.filter(date__gte=start_date, date__lte=end_date, proxy__active=True).order_by('date', 'id')

    if car_id > 0:
        item_tracking_list = item_tracking_list.filter(item__car_id=car_id)
        proxy_tracking_list = proxy_tracking_list.filter(proxy__car_id=car_id)

    item_tracking = []
    for track in item_tracking_list:
        if track.owner.id == driver_id or driver_id == 0:
            item_tracking.append(track)
        elif track.last_owner():
            if track.last_owner().id == driver_id:
                item_tracking.append(track)
    proxy_tracking = []
    for track in proxy_tracking_list:
        if track.owner.id == driver_id or driver_id == 0:
            proxy_tracking.append(track)
        elif track.last_owner():
            if track.last_owner().id == driver_id:
                proxy_tracking.append(track)

    if not xls:
        context = {
            'item_tracking': item_tracking,
            'proxy_tracking': proxy_tracking
        }
        return render(request, 'td/reports/tracking.html', context)
    else:
        from main import settings
        import openpyxl
        from openpyxl.styles import Border, Side, PatternFill, Alignment
        import os
        from openpyxl.writer.excel import save_virtual_workbook

        wb = openpyxl.load_workbook(os.path.join(settings.TEMPLATES[0]['DIRS'][0], 'td', 'reports', 'tracking.xlsx'))
        ws = wb.active

        thin = Side(border_style="thin", color="000000")
        #border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="faf2cc")
        align = Alignment(horizontal="center", vertical="center")

        row = 5
        ws.cell(column=1, row=row, value='Документы/ключи/пропуска')
        ws.merge_cells(start_column=1, end_column=6, start_row=row, end_row=row)
        #ws.cell(column=1, row=row).border = border
        ws.cell(column=1, row=row).fill = fill
        ws.cell(column=1, row=row).alignment = align
        row += 1

        for t in item_tracking:
            ws.cell(column=1, row=row, value=t.date.strftime('%d.%m.%Y'))
            ws.cell(column=2, row=row, value=t.item.get_type_display())
            ws.cell(column=3, row=row, value=str(t.item.car))
            ws.cell(column=4, row=row, value=str(t.item.number))
            if t.last_owner():
                ws.cell(column=5, row=row, value=str(t.last_owner()))
            ws.cell(column=6, row=row, value=str(t.owner))
            row += 1

        ws.cell(column=1, row=row, value='Доверенности')
        ws.merge_cells(start_column=1, end_column=6, start_row=row, end_row=row)
        #ws.cell(column=1, row=row).border = border
        ws.cell(column=1, row=row).fill = fill
        ws.cell(column=1, row=row).alignment = align
        row += 1

        for t in proxy_tracking:
            ws.cell(column=1, row=row, value=t.date.strftime('%d.%m.%Y'))
            ws.cell(column=2, row=row, value=t.proxy.get_type_display())
            ws.cell(column=3, row=row, value=str(t.proxy.car))
            if t.last_owner():
                ws.cell(column=5, row=row, value=str(t.last_owner()))
            ws.cell(column=6, row=row, value=str(t.owner))
            row += 1

        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=report.xlsx"

        return response
