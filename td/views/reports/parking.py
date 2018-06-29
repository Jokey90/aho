from django.shortcuts import render, redirect
from django.http import HttpResponse


def parking_report(request, xls=False):
    from td.models import Parking

    parking_list = Parking.objects.order_by('number').all()

    from main import settings
    import os
    from datetime import date
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.writer.excel import save_virtual_workbook

    wb = openpyxl.load_workbook(os.path.join(settings.TEMPLATES[0]['DIRS'][0], 'td', 'reports', 'parking.xlsx'))
    ws = wb.active
    wrap_alignment = Alignment(wrap_text=True)

    row = 5
    ws.cell(column=5, row=2, value=date.today().strftime('%d.%m.%Y'))
    for n, p in enumerate(parking_list):
        ws.cell(column=1, row=row, value=n+1).style = 'cell'
        ws.cell(column=2, row=row, value=p.number).style = 'cell'
        ws.cell(column=3, row=row, value=p.floor_number).style = 'cell'
        ws.cell(column=4, row=row, value=p.car_name).style = 'cell'
        if p.owner:
            ws.cell(column=5, row=row, value=str(p.owner)).style = 'cell'
        else:
            ws.cell(column=5, row=row).style = 'cell'
        ws.cell(column=6, row=row, value=p.grade).style = 'cell'
        ws.cell(column=7, row=row, value=p.get_payment_type_display()).style = 'cell'
        ws.cell(column=8, row=row, value=p.get_basis_display()).style = 'cell'
        ws.cell(column=9, row=row, value=p.comment).style = 'cell'
        row += 1

    response = HttpResponse(
        save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=report.xlsx"

    return response
