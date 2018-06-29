from django.shortcuts import render, redirect
from django.http import HttpResponse


def to_report_form(request):
    from td.forms import CarPeriodForm
    from datetime import date

    form = CarPeriodForm()

    context = {
        'form': form
    }
    return render(request, 'td/reports/to_form.html', context)


def to_report(request, xls=False):
    from td.models import TO
    from td.forms import CarPeriodForm
    from datetime import datetime

    if request.method != 'POST':
        return redirect(to='td:to_report_form')

    form = CarPeriodForm(request.POST)
    car_id = int(form.data['car'])
    if form.data['start_date']:
        start_date = datetime.strptime(form.data['start_date'], '%d.%m.%Y')
    else:
        start_date = datetime(2000, 1, 1, 0, 0)
    if form.data['end_date']:
        end_date = datetime.strptime(form.data['end_date'], '%d.%m.%Y')
    else:
        end_date = datetime(2099, 12, 31, 0, 0)

    if car_id > 0:
        to_list = TO.objects.filter(car_id=car_id, date__gte=start_date, date__lte=end_date).order_by('date')
    else:
        to_list = TO.objects.filter(date__gte=start_date, date__lte=end_date).order_by('date')

    budget_total = 0
    fact_total = 0

    for to in to_list:
        budget_total += to.budget_amount
        fact_total += to.fact_amount

    if not xls:

        context = {
            'to_list': to_list,
            'budget_total': budget_total,
            'fact_total': fact_total
        }
        return render(request, 'td/reports/to.html', context)
    else:
        from main import settings
        import openpyxl
        from openpyxl.styles import Alignment
        import os
        from openpyxl.writer.excel import save_virtual_workbook

        wb = openpyxl.load_workbook(os.path.join(settings.TEMPLATES[0]['DIRS'][0], 'td', 'reports', 'to.xlsx'))
        ws = wb.active
        wrap_alignment = Alignment(wrap_text=True)

        row = 5
        for to in to_list:
            ws.cell(column=1, row=row, value=to.car.name+' '+to.car.number).style = 'cell'
            ws.cell(column=1, row=row).alignment = wrap_alignment
            ws.cell(column=2, row=row, value=to.date.strftime('%d.%m.%Y')).style = 'cell'
            ws.cell(column=3, row=row, value=to.name).style = 'cell'
            ws.cell(column=4, row=row, value=to.budget_amount).style = 'cell'
            ws.cell(column=5, row=row, value=to.fact_amount).style = 'cell'
            ws.cell(column=6, row=row, value=to.comment).style = 'cell'
            ws.cell(column=6, row=row).alignment = wrap_alignment
            row += 1

        ws.cell(column=1, row=row, value='Итого:').style = 'totals'
        ws.cell(column=2, row=row).style = 'totals'
        ws.cell(column=3, row=row).style = 'totals'
        ws.cell(column=4, row=row, value=budget_total).style = 'totals'
        ws.cell(column=5, row=row, value=fact_total).style = 'totals'
        ws.cell(column=6, row=row).style = 'totals'

        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=report.xlsx"

        return response
