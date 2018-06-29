from django.shortcuts import render
from django.http import HttpResponse


def general_report(request, xls=False):
    from td.models import Car, TO, Mileage, Item, Proxy

    cars = []
    for car in Car.objects.filter(active=True).order_by('name', 'number'):
        cars.append({
            'car': car,
            'to': TO.objects.filter(car=car, fact_amount__gt=0).order_by('-date').first(),
            'mileage': Mileage.objects.filter(car=car).order_by('-year', '-month').first(),
            'items': Item.objects.filter(car=car, active=True).order_by('type', 'end_date'),
            'proxies': Proxy.objects.filter(car=car, active=True).order_by('type'),
        })

    if not xls:
        context = {'cars': cars}
        return render(request, 'td/reports/general.html', context)
    else:
        from main import settings
        from datetime import date
        import openpyxl
        import os
        from openpyxl.writer.excel import save_virtual_workbook

        wb = openpyxl.load_workbook(os.path.join(settings.TEMPLATES[0]['DIRS'][0], 'td', 'reports', 'general.xlsx'))
        ws = wb.active
        row = 2

        for n, car in enumerate(cars):
            drivers = ''
            fuel_card = ''
            osago = ''
            kasko = ''
            pts = ''
            sts = ''

            drivers_state = 'ok'
            for proxy in car['proxies']:
                if proxy.type == 0 and proxy.end_date >= date.today():
                    if drivers != '':
                        drivers += ', '
                    drivers += str(proxy.driver)
                    if (proxy.end_date - date.today()).days < 30:
                        drivers_state = 'warning'

            osago_state = 'ok'
            kasko_state = 'ok'
            for item in car['items']:
                if item.type == 'fuel_card':
                    fuel_card = item.number
                if item.type == 'osago':
                    osago = item.end_date.strftime('%d.%m.%Y')
                    if (item.end_date - date.today()).days < 30:
                        osago_state = 'warning'
                if item.type == 'kasko':
                    kasko = item.end_date.strftime('%d.%m.%Y')
                    if (item.end_date - date.today()).days < 30:
                        kasko_state = 'warning'
                if item.type == 'pts':
                    pts = item.number
                if item.type == 'sts':
                    sts = item.number

            ws.cell(column=1, row=row, value=n + 1).style = 'cell'
            ws.cell(column=2, row=row, value=car['car'].name).style = 'cell'
            ws.cell(column=3, row=row, value=car['car'].number).style = 'cell'
            ws.cell(column=4, row=row, value=drivers).style = 'cell'
            if drivers_state == 'warning' or drivers == '':
                ws.cell(column=4, row=row).style = 'cell-warning'
            ws.cell(column=5, row=row, value=str(car['car'].owner or '')).style = 'cell'
            ws.cell(column=6, row=row, value=fuel_card).style = 'cell'
            ws.cell(column=7, row=row, value=osago).style = 'cell'
            if osago_state == 'warning':
                ws.cell(column=7, row=row).style = 'cell-warning'
            ws.cell(column=8, row=row, value=kasko).style = 'cell'
            if kasko_state == 'warning':
                ws.cell(column=8, row=row).style = 'cell-warning'
            ws.cell(column=9, row=row, value=pts).style = 'cell'
            ws.cell(column=10, row=row, value=sts).style = 'cell'
            ws.cell(column=11, row=row, value=str(car['mileage'].value) + ' км').style = 'cell'

            row += 1

        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=report.xlsx"

        return response

