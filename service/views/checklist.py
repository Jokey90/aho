from django.shortcuts import HttpResponse, redirect, get_object_or_404
from main.backends import render_mobile as render
from django.contrib import messages


def checklist_new(request):
    from django.forms import formset_factory
    from service.forms import ChecklistNewForm, ChecklistValueForm
    from service.models import ChecklistRow, ChecklistValue, Checklist

    checklist_form = ChecklistNewForm()
    checklist_rows = [x for x in ChecklistRow.objects.order_by('zone__ord_number', 'ord_number').all().prefetch_related('zone')]
    rows_num = len(checklist_rows)
    rows_factory = formset_factory(ChecklistValueForm, extra=rows_num)

    if request.method == 'GET':

        formset = rows_factory()
        rows = []

        for i in range(rows_num):
            formset[i].initial = {'row': checklist_rows[i].id}
            rows.append({
                'zone': checklist_rows[i].zone.name,
                'name': checklist_rows[i].name,
                'form': formset[i]
            })

        context = {
            'checklist_form': checklist_form,
            'formset': rows,
            'management_form': formset.management_form
        }
        return render(request, 'service/checklist/new.html', context)

    elif request.method == 'POST':

        form = ChecklistNewForm(request.POST)
        formset = rows_factory(request.POST, request.FILES)
        if formset.is_valid() and form.is_valid():
            checklist = form.save()
            for row in formset:
                row_value = row.save(commit=False)
                row_value.checklist = checklist
                row_value.save()
            messages.success(request, 'Отчет создан')
            return redirect(to='service:checklist_list')
        else:
            messages.error(request, 'Есть ошибки в заполнении формы:<br/>'+str(formset.errors))
            return redirect(to='service:checklist_new')


def checklist_list(request, state='all'):
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from service.models import Checklist

    page = request.GET.get('page', 1)

    if state == 'all':
        data = Checklist.objects.order_by('-date')
        title = 'Отчеты'
    elif state == 'issues':
        data = Checklist.objects.unsolved_only().order_by('-date')
        title = 'Отчеты с замечаниями'
    else:
        data = Checklist.objects.solved_only().order_by('-date')
        title = 'Отчеты без замечаний'

    paginator = Paginator(data, 25)

    try:
        checklists = paginator.page(page)
    except PageNotAnInteger:
        checklists = paginator.page(1)
    except EmptyPage:
        checklists = paginator.page(paginator.num_pages)

    context = {
        'checklists': checklists,
        'title': title
    }
    return render(request, 'service/checklist/list.html', context)


def checklist_info(request, uid):
    from service.models import Checklist, Zone

    checklist = get_object_or_404(Checklist, id=uid)

    checklist_zones = []
    for zone in Zone.objects.order_by('ord_number'):
        checklist_zones.append({
            'name': zone.name,
            'issues': checklist.checklistvalue_set.filter(row__zone=zone, has_issues=True).order_by('row__ord_number'),
            'unsolved_count': checklist.checklistvalue_set.filter(row__zone=zone, has_issues=True, solve_date__isnull=True).count()
        })

    context = {
        'checklist': checklist,
        'checklist_zones': checklist_zones
    }

    return render(request, 'service/checklist/info.html', context)


def checklist_edit(request, uid):
    from django.forms import modelformset_factory
    from service.models import Checklist, ChecklistValue
    from service.forms import ChecklistEditForm, ChecklistValueEditInlineForm

    checklist = get_object_or_404(Checklist, id=uid)

    queryset = ChecklistValue.objects.filter(checklist_id=uid).order_by('row__zone__ord_number','row__ord_number').prefetch_related('row', 'row__zone')
    rows_num = queryset.count()
    rows_factory = modelformset_factory(ChecklistValue, form=ChecklistValueEditInlineForm, extra=rows_num)

    if request.method == 'GET':

        checklist_form = ChecklistEditForm(instance=checklist)
        formset = rows_factory(queryset=queryset)

        rows = []

        for i, elem in enumerate(queryset):
            rows.append({
                'zone': elem.row.zone.name,
                'name': elem.row.name,
                'form': formset[i]
            })

        context = {
            'checklist_form': checklist_form,
            'formset': rows,
            'management_form': formset.management_form
        }
        return render(request, 'service/checklist/edit.html', context)

    elif request.method == 'POST':

        checklist_form = ChecklistEditForm(request.POST, instance=checklist)
        formset = rows_factory(request.POST, request.FILES, queryset=queryset)

        if checklist_form.is_valid() and formset.is_valid():
            checklist_form.save()
            instances = formset.save(commit=False)
            for instance in instances:
                instance.save()
            messages.success(request, 'Изменения применены')
            return redirect(to='service:checklist_info', uid=uid)
        else:
            rows = []

            for i, elem in enumerate(queryset):
                rows.append({
                    'zone': elem.row.zone.name,
                    'name': elem.row.name,
                    'form': formset[i]
                })

            context = {
                'checklist_form': checklist_form,
                'formset': rows,
                'management_form': formset.management_form
            }
            return render(request, 'service/checklist/edit.html', context)


def checklist_xls(request, uid):
    from service.models import Checklist, Zone

    checklist = get_object_or_404(Checklist, id=uid)

    from main import settings
    from datetime import date
    import openpyxl
    import os
    from openpyxl.writer.excel import save_virtual_workbook

    wb = openpyxl.load_workbook(os.path.join(settings.TEMPLATES[0]['DIRS'][0], 'service', 'checklist', 'report.xlsx'))
    ws = wb.active

    ws.oddHeader.center.text = 'Дата обхода: ' + checklist.date.strftime('%d.%m.%Y')

    ws.merge_cells(start_column=1, end_column=6, start_row=1, end_row=1)
    ws.cell(column=1, row=1, value='Чек лист обхода помещений '+checklist.floor.name+' эт.').style = 'header1'

    row = 4

    for zone_number, zone in enumerate(Zone.objects.order_by('ord_number')):
        ws.merge_cells(start_column=1, end_column=6, start_row=row, end_row=row)
        ws.cell(column=1, row=row, value=str(zone_number+1)+'. '+zone.name).style = 'header1'
        row += 1

        for val_num, value in enumerate(checklist.checklistvalue_set.filter(row__zone=zone).order_by('row__ord_number')):
            ws.cell(column=1, row=row, value=str(zone_number+1)+'.'+str(val_num+1)).style = 'cell'
            ws.cell(column=2, row=row, value=value.row.name).style = 'cell'
            if len(value.row.name)>30:
                ws.row_dimensions[row].height *= 2
            if value.has_issues:
                ws.cell(column=3, row=row, value='-').style = 'cell'
            else:
                ws.cell(column=3, row=row, value='+').style = 'cell'
            if value.solve_date is not None:
                ws.cell(column=4, row=row, value='да').style = 'cell'
                ws.cell(column=5, row=row, value=value.solve_date.strftime('%d.%m.%Y')).style = 'cell'
            else:
                ws.cell(column=4, row=row, value='').style = 'cell'
                ws.cell(column=5, row=row, value='').style = 'cell'
            ws.cell(column=6, row=row, value=value.comment).style = 'cell'
            row += 1

    row += 2
    ws.cell(column=2, row=row, value='___________________________________________')
    ws.cell(column=5, row=row, value='____________________________')
    row += 1
    ws.cell(column=2, row=row, value='Директор по управлению персоналом (подпись)')
    ws.cell(column=5, row=row, value='Обход провел Ф.И.О. (подпись)')

    response = HttpResponse(
        save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=report.xlsx"

    return response
